"""Equipment bulk import: parse → validate (dry run) → import.

Pure functions; no HTTP concerns. Views feed files in and render the results.
"""

import csv
import io
from dataclasses import dataclass, field
from datetime import date, datetime

from .models import Department, Equipment


class ImportFormatError(Exception):
    pass


def parse_upload(file_obj, filename) -> list[dict]:
    """Return one {header: cell} dict per data row. Headers lower-cased and
    stripped; values stringified and stripped ("" for empty cells)."""
    name = filename.lower()
    if name.endswith(".csv"):
        raw = file_obj.read().decode("utf-8-sig")
        table = list(csv.reader(io.StringIO(raw)))
    elif name.endswith(".xlsx"):
        from openpyxl import load_workbook

        ws = load_workbook(file_obj, read_only=True, data_only=True).active
        table = [list(row) for row in ws.iter_rows(values_only=True)]
    else:
        raise ImportFormatError("Only .csv and .xlsx files are supported.")

    table = [row for row in table if any(c not in (None, "") for c in row)]
    if len(table) < 2:
        raise ImportFormatError("File needs a header row and at least one data row.")

    headers = [str(h or "").strip().lower() for h in table[0]]
    rows = []
    for raw_row in table[1:]:
        cells = ["" if c is None else str(c).strip() for c in raw_row]
        row_dict = dict(zip(headers, cells[:len(headers)]))
        # Add extra cells beyond header count with synthetic keys (column_N)
        for i, cell in enumerate(cells[len(headers):], start=len(headers) + 1):
            if cell:  # Only add non-empty extra cells
                row_dict[f"column_{i}"] = cell
        rows.append(row_dict)
    return rows


REQUIRED_COLUMNS = ("name", "serial_number", "department")
OPTIONAL_COLUMNS = (
    "manufacturer",
    "vendor",
    "model_number",
    "purchase_date",
    "installation_date",
    "is_critical_asset",
)
DATE_COLUMNS = ("purchase_date", "installation_date")
TRUTHY = {"true", "yes", "1"}


@dataclass
class RowResult:
    line: int
    data: dict = field(default_factory=dict)
    department_name: str = ""
    extra: dict = field(default_factory=dict)
    errors: list = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return not self.errors


def _parse_date(value: str) -> date:
    return datetime.strptime(value, "%Y-%m-%d").date()


def validate_rows(rows, create_missing_departments=False) -> list[RowResult]:
    """Dry run: per-row cleaning + errors. Never touches the database beyond
    reads. Line numbers are 2-based to match the spreadsheet."""
    existing_serials = set(
        Equipment.objects.values_list("serial_number", flat=True)
    )
    known_departments = set(Department.objects.values_list("name", flat=True))
    seen_serials: set[str] = set()
    results = []
    for index, row in enumerate(rows):
        result = RowResult(line=index + 2)
        for column in REQUIRED_COLUMNS:
            if not row.get(column, ""):
                result.errors.append(f"missing required column: {column}")
        serial = row.get("serial_number", "")
        if serial:
            if serial in seen_serials:
                result.errors.append(f"duplicate serial_number in file: {serial}")
            elif serial in existing_serials:
                result.errors.append(f"serial_number already exists: {serial}")
            seen_serials.add(serial)
        department = row.get("department", "")
        if (
            department
            and department not in known_departments
            and not create_missing_departments
        ):
            result.errors.append(f"unknown department: {department}")
        result.department_name = department

        for key, value in row.items():
            if key == "department":
                continue  # resolved to a Department FK at import time
            if key in DATE_COLUMNS:
                if value:
                    try:
                        result.data[key] = _parse_date(value)
                    except ValueError:
                        result.errors.append(f"bad {key} (want YYYY-MM-DD): {value}")
            elif key == "is_critical_asset":
                result.data[key] = value.lower() in TRUTHY
            elif key in REQUIRED_COLUMNS or key in OPTIONAL_COLUMNS:
                if value:
                    result.data[key] = value
            elif value:
                result.extra[key] = value
        results.append(result)
    return results
